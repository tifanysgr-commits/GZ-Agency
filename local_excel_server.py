from __future__ import annotations

import json
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook


HOST = "127.0.0.1"
PORT = 8765
PROJECT_ROOT = Path(__file__).resolve().parent
EXCEL_PATH = PROJECT_ROOT / "respuestas.xlsx"
HEADERS = [
    "nombre",
    "email",
    "mensaje",
    "fecha_envio",
    "negocio",
    "telefono",
    "tipo_app",
    "fase_negocio",
    "usuarios_clientes",
    "factura_actualmente",
    "origen",
]


def ensure_workbook() -> None:
    if EXCEL_PATH.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Respuestas"
    sheet.append(HEADERS)
    workbook.save(EXCEL_PATH)


def append_response(payload: dict) -> None:
    ensure_workbook()

    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active
    row = [
        payload.get("nombre", ""),
        payload.get("email", ""),
        payload.get("mensaje", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        payload.get("negocio", ""),
        payload.get("telefono", ""),
        payload.get("tipo_app", ""),
        payload.get("fase_negocio", ""),
        payload.get("usuarios_clientes", ""),
        payload.get("factura_actualmente", ""),
        payload.get("origen", ""),
    ]
    sheet.append(row)
    workbook.save(EXCEL_PATH)


class LocalExcelHandler(BaseHTTPRequestHandler):
    def _send_json(self, status_code: int, body: dict) -> None:
        response = json.dumps(body).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(response)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(response)

    def do_OPTIONS(self) -> None:
        self._send_json(200, {"ok": True})

    def do_POST(self) -> None:
        if self.path != "/api/form-responses":
            self._send_json(404, {"ok": False, "error": "Not found"})
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length)

        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json(400, {"ok": False, "error": "Invalid JSON"})
            return

        if not payload.get("nombre") or not payload.get("email") or not payload.get("mensaje"):
            self._send_json(400, {"ok": False, "error": "Missing required fields"})
            return

        try:
            append_response(payload)
        except Exception as exc:  # pragma: no cover
            self._send_json(500, {"ok": False, "error": str(exc)})
            return

        self._send_json(200, {"ok": True, "file": str(EXCEL_PATH.name)})

    def log_message(self, format: str, *args) -> None:
        return


if __name__ == "__main__":
    ensure_workbook()
    server = HTTPServer((HOST, PORT), LocalExcelHandler)
    print(f"Local Excel server running at http://{HOST}:{PORT}")
    print(f"Saving responses to {EXCEL_PATH}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
